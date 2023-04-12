""" main file - Run this should prompt user for a properly formatted Excel
file then update that file with pretty output of championship standings"""

import tkinter
import tkinter.filedialog
import openpyxl
import openpyxl.worksheet
import openpyxl.utils
import openpyxl.styles
from openpyxl.cell.cell import Cell

# Graphic Range Type Hint
GraphicRange = tuple[tuple[Cell, ...], ...]

# Bold Gold for all counted 1st place finishes - Bold Silver second, Bold-Bronze third, Bold- Black - counted
# non-podium finishes
boldGold = openpyxl.styles.Font(bold=True, color="ffc200")
BoldSilver = openpyxl.styles.Font(bold=True, color="9a9a9a")
BoldBronze = openpyxl.styles.Font(bold=True, color="CD7F32")
BoldBlack = openpyxl.styles.Font(bold=True)

# Dull grey without bold for non-counting (dropped) finishes
DroppedGrey = openpyxl.styles.Font(color="e6e6e6")

# Assign the appropriate styles to the finishes (both str and int forms)
STYLES = {"1st": boldGold, "2nd": BoldSilver, "3rd": BoldBronze, 1: boldGold, 2: BoldSilver, 3: BoldBronze}

# Points dictionary Finish(int): points for that finish(int)
POINTS = {1: 25, 2: 18, 3: 15, 4: 12, 5: 10, 6: 8, 7: 6, 8: 4, 9: 2, 10: 1}

# CONFIG PARAMETER - Number of non-counted weeks in a season.
DROPPED_WEEKS = 2


def get_wb() -> openpyxl.Workbook:
    """
    Launches a tkinter file selection dialog - Returns a workbook object else raises an error if there are issues.

    Also prints a simple error message to log.
    User Cancel or blank selection -> FileNotFoundError
    User selects invalid file openpyxl can't handle -> NameError

    :return: a valid workbook object for use
    :raises FileNotFoundError: if the user clicks cancel
    :raises NameError: if the user selects a file openpyxl can't parse
    """
    # this launches a generic tkinter file open dialog in the pwd
    root_window = tkinter.Tk()
    root_window.withdraw()
    file_path = tkinter.filedialog.askopenfilename()

    # this will generally trigger if user clicked cancel
    try:
        assert file_path != ''
    except AssertionError:
        print("You did not select a valid file.")
        raise FileNotFoundError

    # This will trigger if the user selected a bad file tha openpyxl can't parse
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("error opening workbook")
        print(e)
        raise NameError

    return wb


class Series(object):
    def __init__(self, sheet, drop_weeks):
        self.sheet = sheet
        self.drivers = []
        self.weeks = []
        self.drop_weeks = drop_weeks
        self.weekly_sorted_drivers = []

    def runcalc(self):
        self.read_weeks()
        self.read_drivers()
        for driver in self.drivers:
            for week in self.weeks:
                driver.calc_points_results_by_week(week, self.drop_weeks)
        for week in self.weeks:
            self.sort_drivers(week)

    def read_drivers(self):
        i = 3
        while True:
            cell_value = self.sheet.cell(row=1, column=i).value
            if cell_value is None:
                break
            d = Driver(cell_value, i)
            assert len(self.weeks) != 0
            d.read_results(self.sheet, len(self.weeks))
            self.drivers.append(d)
            i += 1

    def read_weeks(self):
        i = 2
        while True:
            cell_value = self.sheet.cell(row=i, column=1).value
            if cell_value is None:
                break
            self.weeks.append(cell_value)
            i += 1

    def sort_drivers(self, week):
        sorted_drivers = self.drivers.copy()
        # sort by number of points for the week. If that is not enough then sort based on best dropped position
        # going through all dropped positions as needed. - Need to still add Alpha drivers name last?
        sorted_drivers = sorted(sorted_drivers,
                                key=lambda y: [-y.weekly_points[week - 1]] + [y.weekly_dropped[week - 1][p][1] for p in
                                                                              range(len(y.weekly_dropped[week - 1]))])
        # currently sorting on DN* = 10000, May need alphabetical drivers name added as final condition
        # or could make these position object then define the __lt__ for position class?
        self.weekly_sorted_drivers.append(sorted_drivers)

    def get_num_weeks(self):
        return len(self.weeks)

    def get_num_drivers(self):
        return len(self.drivers)


class Driver(object):
    def __init__(self, name, col):
        self.name = name
        self.col = col
        self.all_results = []
        self.this_week_results = []
        self.weekly_results = []
        self.weekly_points = []
        self.weekly_dropped = []

    def read_results(self, sheet, num_weeks):
        for i in range(2, num_weeks + 2):
            self.all_results.append((i - 1, sheet.cell(row=i, column=self.col).value))

    def calc_points_results_by_week(self, num_weeks, drop_weeks):
        week_results = self.all_results[0:num_weeks]
        dropped_results = []

        for i in range(len(week_results)):
            if week_results[i][1] in POINTS.keys():
                week_results[i] = (week_results[i][0], week_results[i][1], POINTS[week_results[i][1]])
            else:
                week_results[i] = (week_results[i][0], week_results[i][1], 0)

        week_results = sorted(week_results, key=lambda y: y[2], reverse=True)

        if num_weeks > drop_weeks:
            for i in range(len(week_results)):
                if i < (num_weeks - drop_weeks):
                    week_results[i] = (week_results[i][0], week_results[i][1], week_results[i][2], 1)
                else:
                    week_results[i] = (week_results[i][0], week_results[i][1], week_results[i][2], 0)
                    dropped_results.append((week_results[i][0], week_results[i][1], week_results[i][2], 0))
        else:
            week_results[0] = (week_results[0][0], week_results[0][1], week_results[0][2], 1)
            for i in range(1, len(week_results)):
                week_results[i] = (week_results[i][0], week_results[i][1], week_results[i][2], 0)
                dropped_results.append((week_results[i][0], week_results[i][1], week_results[i][2], 0))

        # Tuple format at this point is (week,position,points for position,dropped?)
        # make it (week,position_numerical,position for print,points for position, dropped)

        self.create_print_positions(week_results)
        self.create_print_positions(dropped_results)

        # sort all the dropped results in order of quality as we may need them to break ties.
        dropped_results = sorted(dropped_results, key=lambda y: y[1])
        self.weekly_results.append(week_results)
        self.weekly_dropped.append(dropped_results)
        score = 0
        for item in week_results:
            if item[4] == 1:
                score += item[3]
        self.weekly_points.append(score)

    def create_print_positions(self, week_results):
        """ take in list of Tuples format at this point is (week,position,points for position,dropped?)
        make it list of tuples (week,position_numerical,position print,points for position, dropped)"""
        endings_dict = {1: "st", 2: "nd", 3: "rd"}
        for i in range(len(week_results)):
            if type(week_results[i][1]) == str:
                week_results[i] = (week_results[i][0], 1000, week_results[i][1], week_results[i][2], week_results[i][3])
            elif week_results[i][1] in endings_dict.keys():
                week_results[i] = (
                    week_results[i][0], week_results[i][1], str(week_results[i][1]) + endings_dict[week_results[i][1]],
                    week_results[i][2], week_results[i][3])
            else:
                week_results[i] = (week_results[i][0], week_results[i][1], str(week_results[i][1]) + 'th',
                                   week_results[i][2], week_results[i][3])


class ChampWorkBook(object):
    """
    Championship wookbook object - An onpenpyxl workbook representing several championships

    Holds a lot of formatting methods and the raw excel data
    """

    def __init__(self, wb: openpyxl.Workbook, raw_str: str, summary_str: str) -> None:
        """
        Initiate the Champ Workbook object.

        Creates lists of the raw input sheets - Summary Sheets and other sheets to ignor and copy across.
        Creates a blank holder for series objects.

        :param wb: A valid openpyxl workbook containing championship data matching the example format
        :param raw_str: A string that precedes the series name in each tab full of input date in the input excel doc.
        :param summary_str: A string that will proceed each output tab name followed by the series name.
        """
        self.wb = wb
        self.raw_data_sheets = []
        self.raw_str = raw_str
        self.summary_sheets = []
        self.summary_str = summary_str
        # other sheets are sheets that will be copied across unaltered and otherwise ignored
        self.others_sheets = []
        # blank list for storing series objects
        self.series = []

        # figure out which sheet objects are of which type then add them to their appropriate lists
        l_raw = len(raw_str)
        l_summary = len(summary_str)
        for sheet in wb.sheetnames:
            if sheet[:l_raw] == raw_str:
                self.raw_data_sheets.append(sheet)
            elif sheet[:l_summary] == summary_str:
                self.summary_sheets.append(sheet)
            else:
                self.others_sheets.append(sheet)

        # Wipe old summary sheets since they will be generated anew
        for sheet in self.summary_sheets:
            self.wb.remove(self.wb[sheet])
        self.summary_sheets = []

    def calc_series(self, drop_weeks: int) -> None:
        """
        Run the calculation of all the basic series represented by the raw data sheets

        Adds a series object based off of each sheet to the series list
        then for each series in the list - tells that series to perform its calculation method.

        :param drop_weeks: 0 or positive int for the number of weeks in a season to be dropped
        :return: None - Alters State of the parent.
        """

        # Create a series object in the self series list for each raw sheet
        for sheet_name in self.raw_data_sheets:
            sheet = self.wb[sheet_name]
            self.series.append(Series(sheet, drop_weeks))

        # run built in calculation on each of those series
        for series in self.series:
            series.runcalc()

    def init_out_sheets(self) -> None:
        """
        create any missing blank summary sheets then format all sheets grid framework based on drivers and weeks raced
        so far. Finally print out all results data.s

        :return: Alters state - New results sheets created
        """

        # check each raw sheet - create new summary sheet name -If that sheet doesn't already exist create it for real
        for sheet in self.raw_data_sheets:
            new_summary = sheet.replace(self.raw_str, self.summary_str)
            if new_summary not in self.summary_sheets:
                self.wb.create_sheet(new_summary)
                self.summary_sheets.append(new_summary)

        # run each sheet through the grid printing functions, so they have all their combined cells and grids set.
        # then write in the calculated results data
        for i in range(len(self.series)):
            self.format_out_sheet(self.summary_sheets[i], self.series[i].get_num_weeks(),
                                  self.series[i].get_num_drivers())
            self.write_data(self.summary_sheets[i], self.series[i])

    def calc_graphic_range(self, sheet: str, num_racers: int, num_weeks: int) -> GraphicRange:
        """
        Returns a graphics range for the cells covering the expected display output
        :param sheet: str of a sheet name
        :param num_racers: int of number of racers
        :param num_weeks: int of number of weeks
        :return: GraphicRange - tuple of tuples of cells
        """
        # 3 lines worth of headers
        output_height = 3 + num_racers
        # 2X num weeks covers driver and pts column each week then need to add positions
        output_width = 2 * num_weeks
        for i in range(num_weeks):
            output_width += i + 1
        ws = self.wb[sheet]
        end_letter = openpyxl.utils.cell.get_column_letter(output_width)
        end_cell = end_letter + str(output_height)
        graphic_range = ws['A1':end_cell]
        return graphic_range

    def format_out_sheet(self, sheet: str, num_weeks: int, num_racers: int) -> None:
        """
        This routine alters the sheets in the wb to have the appropriate grid lines and merged cells for all weeks and
        all the basic headers and labels - Not drivers names/points or finishes.

        :param sheet: a string representing a valid summary worksheet tab in the self.wb openpyxl Workbook object
        :param num_weeks: The number of completed weeks in the corresponding series - Integers 1+
        :param num_racers:  The number of competitors in the series - Integers 1+
        """

        # 3 lines worth of headers
        output_height = 3 + num_racers
        # 2X num weeks covers driver and pts column each week then need to add positions
        output_width = 2 * num_weeks
        # add 1 column for week 1, 2 for week 2 etc.
        for i in range(num_weeks):
            output_width += i + 1
        ws = self.wb[sheet]
        # create the top header cell full width
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=output_width)

        # create all the week header merged cells( cols 1-3 for week 1, 4-7 for week2 etc)
        start_week_col = 1
        for i in range(1, num_weeks + 1):
            ws.merge_cells(start_row=2, start_column=start_week_col, end_row=2, end_column=start_week_col + i + 1)
            if i != 1:
                ws.merge_cells(start_row=3, start_column=start_week_col + 2, end_row=3,
                               end_column=start_week_col + 1 + i)
            start_week_col += i + 2

        # Get bottom right cell then create graphic range from top left to bottom right
        end_letter = openpyxl.utils.cell.get_column_letter(output_width)
        end_cell = end_letter + str(output_height)
        graphic_range = ws['A1':end_cell]

        # center and white background fill all cells
        white_fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                                 fill_type='solid')  # change to FFFFFF
        center_all = openpyxl.styles.alignment.Alignment(horizontal="center", vertical="center")
        for row in graphic_range:
            for cell in row:
                cell.alignment = center_all
                cell.fill = white_fill

        # setup border styles
        thick_black = openpyxl.styles.Side(border_style='medium', color="000000")
        thin_black = openpyxl.styles.Side(border_style='thin', color="000000")
        thick_all_sides_border = openpyxl.styles.Border(left=thick_black, right=thick_black, top=thick_black,
                                                        bottom=thick_black)
        # First Two Headers for series name and weeks - Set Borders
        bold14pt = openpyxl.styles.Font(size=14, bold=True)
        bold_only = openpyxl.styles.Font(bold=True)
        for cell in graphic_range[0]:
            cell.border = thick_all_sides_border
            cell.font = bold14pt
        for cell in graphic_range[1]:
            cell.font = bold_only
            cell.border = thick_all_sides_border

        # Calculate indices for week breaks
        i_week_breaks = []
        prev_stop = 0
        for i in range(num_weeks):
            i_week_breaks.append(prev_stop + 2 + i + 1)
            prev_stop += 3 + i

        # Format row 3 with thick upper lines and thick start stop week breaks. Thin lower line
        row3 = graphic_range[2]
        for i in range(len(row3)):
            if i == 0:
                row3[i].border = openpyxl.styles.Border(left=thick_black, top=thick_black, bottom=thin_black)
            elif i == output_width - 1:
                row3[i].border = openpyxl.styles.Border(right=thick_black, top=thick_black, bottom=thin_black)
            elif i in i_week_breaks:
                row3[i].border = openpyxl.styles.Border(left=thick_black, top=thick_black, bottom=thin_black)
            elif i + 1 in i_week_breaks:
                row3[i].border = openpyxl.styles.Border(right=thick_black, top=thick_black, bottom=thin_black)
            else:
                row3[i].border = openpyxl.styles.Border(top=thick_black, bottom=thin_black)

        # format last row ith Thick bottom line and thin top line otherwise Thick week breaks.
        row_last = graphic_range[output_height - 1]
        for i in range(len(row_last)):
            if i == 0:
                row_last[i].border = openpyxl.styles.Border(left=thick_black, top=thin_black, bottom=thick_black)
            elif i == output_width - 1:
                row_last[i].border = openpyxl.styles.Border(right=thick_black, top=thin_black, bottom=thick_black)
            elif i in i_week_breaks:
                row_last[i].border = openpyxl.styles.Border(left=thick_black, top=thin_black, bottom=thick_black)
            elif i + 1 in i_week_breaks:
                row_last[i].border = openpyxl.styles.Border(right=thick_black, top=thin_black, bottom=thick_black)
            else:
                row_last[i].border = openpyxl.styles.Border(top=thin_black, bottom=thick_black)

        # Format all interim lines all thin upper and lower. Thick start, stop, week-break.
        for j in range(3, output_height - 1):
            for i in range(output_width):
                if i == 0:
                    graphic_range[j][i].border = openpyxl.styles.Border(left=thick_black, top=thin_black,
                                                                        bottom=thin_black)
                elif i == output_width - 1:
                    graphic_range[j][i].border = openpyxl.styles.Border(right=thick_black, top=thin_black,
                                                                        bottom=thin_black)
                elif i in i_week_breaks:
                    graphic_range[j][i].border = openpyxl.styles.Border(left=thick_black, top=thin_black,
                                                                        bottom=thin_black)
                elif i + 1 in i_week_breaks:
                    graphic_range[j][i].border = openpyxl.styles.Border(right=thick_black, top=thin_black,
                                                                        bottom=thin_black)
                else:
                    graphic_range[j][i].border = openpyxl.styles.Border(top=thin_black, bottom=thin_black)

        # call routine to write the text into the header cells - Get series name by stripping summary_
        self.write_basic_headers(sheet, num_weeks, num_racers, sheet.replace("summary_", ''))

    def write_basic_headers(self, sheet: str, num_weeks: int, num_racers: int, series_name: str) -> None:
        """
        This routine will write the Series name and the week 1, 2, 3 etc headers into the xls sheet
        :param sheet: string pointing to a valid openpyxl workbook sheet/tab
        :param num_weeks: int of 1+ representing number of weeks in series.
        :param num_racers: int of 1+ representing number of racers participating in the series
        :param series_name: String of the series name for printing out
        """

        graphic_range = self.calc_graphic_range(sheet, num_racers, num_weeks)

        # write the series name in the top left most cell (which is really the big merged-full width centered header)
        graphic_range[0][0].value = series_name

        # set up gold silver and bronze backgrounds for top three in points columns
        gold_fill = openpyxl.styles.PatternFill(start_color='FFD700', end_color='FFD700',
                                                fill_type='solid')
        silver_fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0',
                                                  fill_type='solid')
        bronze_fill = openpyxl.styles.PatternFill(start_color='CD7F32', end_color='CD7F32',
                                                  fill_type='solid')
        week = 1

        # find the top left corner of each week subbox then start the writing in all the headers and background colors
        week00cells = self.calc_00_cells(num_weeks)
        for cell_index in week00cells:
            graphic_range[1][cell_index].value = "Week " + str(week)
            week += 1

            graphic_range[2][cell_index].value = "Driver"
            graphic_range[2][cell_index + 1].value = "Pts"
            graphic_range[2][cell_index + 2].value = "Finishes"
            # 1st,2nd,3rd backgrounds
            graphic_range[3][cell_index].fill = gold_fill
            graphic_range[4][cell_index].fill = silver_fill
            graphic_range[5][cell_index].fill = bronze_fill

    def write_data(self, sheet: str, series: Series) -> None:
        """
        Writes all the driver finish data and points into the appropriate cells
        :param sheet: valid string for an output sheet of the openpyxl workbook
        :param series: A calculated and ready for printing series object
        :return: No return
        """
        ws = self.wb[sheet]

        # get locations for the top left cell that now needs input - Should be the Driver in first each week location
        num_weeks = series.get_num_weeks()
        first_cells_lr = self.calc_00_cells(num_weeks)
        first_cells_vert = 4

        # print week by week then driver by driver
        # TODO: Implement a get_weeks method
        for week in series.weeks:
            index = week - 1
            # set the current working cell to the first place driver's name position for the current week
            curr_cell = ws.cell(row=first_cells_vert, column=first_cells_lr[index] + 1)
            # set driver index to zero for the first driver - this will then increment to move downsheet per driver.
            driver_index = 0
            # Driver position - just the driver index plus 1 since the finish position style list isn't zero indexed
            driver_pos = driver_index+1

            # iterate over this week's drivers in order
            # - TODO: Implement a method in series rather than indexing into and internal structure
            for driver in series.weekly_sorted_drivers[index]:
                # Write Driver's name into cell - TODO: Implement a get_name method on the driver class
                curr_cell.value = driver.name

                # Set color style for pts text then enter the drivers point in that cell (one to right)
                if driver_pos in [1, 2, 3]:
                    style = STYLES[driver_pos]
                else:
                    style = BoldBlack
                points_cell = curr_cell.offset(column=1)
                # TODO: Implement a get points by week method
                points_cell.value = driver.weekly_points[index]
                points_cell.font = style
                print(type(points_cell))

                # call the print positions function to continue printing the positions for that week out to the right.
                self.print_positions(driver, week, points_cell)

                # move on to the next driver selecting the next cell down
                driver_index += 1
                # Driver position - just the driver index plus 1 since the finish position style list isn't zero indexed
                driver_pos = driver_index + 1
                curr_cell = ws.cell(row=first_cells_vert + driver_index, column=first_cells_lr[index] + 1)

    def print_positions(self, driver: Driver, week: int, start_pos: Cell) -> None:
        """
        Prints formatted text of all the drivers finishes for a given week starting one cell right of the passed cell
        :param driver: A valid championship Driver object
        :param week: an int of which week we want results for (human-readable starting at 1 not at 0)
        :param start_pos: openpyxl Cell object just left of where the first position will be printed (The points cell)
        """

        # TODO - Implement a get results by week method
        # get the lists of drivers positions sorted by the week they occurred
        positions = driver.weekly_results[week - 1]
        positions = sorted(positions, key=lambda y: y[0])

        # move one cell right as this is the first place we want to print
        # TODO consider moving this to the code in the funciton that will call this one so this starts in the passed in
        #   cell
        cell_to_write = start_pos.offset(column=1)
        for pos in positions:
            # if the position is a dropped result
            if pos[4] == 0:
                style = DroppedGrey
                cell_to_write.font = style
                cell_to_write.value = pos[2]
            else:
                # if not dropped is it a podium result - if so fancy formatting applied
                if pos[2] in STYLES.keys():
                    style = STYLES[pos[2]]
                    cell_to_write.font = style
                    cell_to_write.value = pos[2]
                else:
                    # Otherwise standard black will be written
                    style = BoldBlack
                    cell_to_write.font = style
                    cell_to_write.value = pos[2]
            # after writing this result keep stepping the cell to the right for next result
            cell_to_write = cell_to_write.offset(column=1)

    def write_out(self):
        default_extension = ".xlsx"
        default_filename = "out" + default_extension
        file_types = [("Excel files", "*" + default_extension), ("All files", "*.*")]
        file_path = tkinter.filedialog.asksaveasfilename(initialfile=default_filename, filetypes=file_types,
                                                         defaultextension=default_extension)
        try:
            assert file_path != ''
        except AssertionError:
            print("You did not select a valid output file.")
            raise FileNotFoundError
        self.wb.save(file_path)

    def calc_00_cells(self, num_weeks: int) -> list[int]:
        """
        Calculate the column for the left of each week sub box
        :param num_weeks: int - number of weeks in the series
        :return: List of ints - Each int is the index for the leftmost entry for that week (the position column)
        """
        week00cells = [0]
        for i in range(1, num_weeks):
            week00cells.append(week00cells[i - 1] + 2 + i)
        return week00cells


if __name__ == '__main__':
    RAW_NAME_BASE = "rawdata_"
    SUMMARY_NAME_BASE = "summary_"

    the_wb = ChampWorkBook(get_wb(), RAW_NAME_BASE, SUMMARY_NAME_BASE)
    the_wb.calc_series(DROPPED_WEEKS)
    the_wb.init_out_sheets()
    the_wb.write_out()
